"""
export_service.py
Exports filtered expenses to a formatted .xlsx file using OpenPyXL.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import fetch_filtered_expenses
from services.expense_service import PAYMENT_LABELS


# ── Colors ────────────────────────────────────────────────────────────────────
_HEADER_BG = "1E3A5F"
_HEADER_FG = "FFFFFF"
_ALT_ROW_BG = "EBF2FF"
_TOTAL_BG = "D0E4FF"
_BORDER_COLOR = "AABBCC"

_THIN = Side(style="thin", color=_BORDER_COLOR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

COLUMNS = ["ID", "Date", "Description", "Category", "Payment Method", "Amount", "Notes"]


def export_to_excel(
    output_path: Path,
    month: str | None = None,
    category: str | None = None,
    payment_method: str | None = None,
) -> Path:
    """
    Build and save a formatted Excel workbook.
    Returns the resolved output path.
    """
    expenses = fetch_filtered_expenses(
        month=month, category=category, payment_method=payment_method
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # ── Title row ─────────────────────────────────────────────────────────────
    title = f"Monthly Expense Tracker — exported {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=13, color=_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row ────────────────────────────────────────────────────────────
    header_row = 2
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=_HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor="2C5282")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    total = 0.0
    for row_idx, expense in enumerate(expenses, start=3):
        is_alt = (row_idx % 2 == 0)
        bg = _ALT_ROW_BG if is_alt else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)

        values = [
            expense.id,
            expense.date,
            expense.description,
            expense.category,
            PAYMENT_LABELS.get(expense.payment_method, expense.payment_method),
            expense.amount,
            expense.notes,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
            if col_idx == 6:  # Amount column
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        total += expense.amount

    # ── Totals row ────────────────────────────────────────────────────────────
    total_row = 3 + len(expenses)
    last_col = len(COLUMNS)

    ws.merge_cells(f"A{total_row}:{get_column_letter(last_col - 2)}{total_row}")
    label_cell = ws.cell(row=total_row, column=1, value="TOTAL")
    label_cell.font = Font(bold=True, size=11)
    label_cell.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    label_cell.alignment = Alignment(horizontal="right", vertical="center")
    label_cell.border = _BORDER

    total_cell = ws.cell(row=total_row, column=last_col - 1, value=total)
    total_cell.font = Font(bold=True, size=11)
    total_cell.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = _BORDER

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [6, 12, 38, 18, 16, 12, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
